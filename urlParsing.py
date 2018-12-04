# python 3.7.1 / 64bit
# pip install beautifulSoup4
# pip install requests
# pip install openpyxl

# *계기
# 사이트의 게시판을 매일 체크할 수 있도록 파싱하여, 
# 게시판의 제목과, 링크를 담은 엑셀로 만들고 싶었습니다.
# 로그인하지 않은 상태에서 접근할 수 있는 범위에서 파싱하였습니다.

# *주요 내용
# URL을 파싱합니다.
# 파싱한 list가 path일때
# path가 한글일때
# exe 파일변환하여 사용하기 위해 console에 출력합니다.
# 트래픽 제한을 방지하기 위해 1초간격으로 일정량만 파싱합니다.
# 파싱한 내역을 제목과 전체 URL을 엑셀로 생성합니다.

# *사용
# 필수 # url을 지정해야합니다.
# 필수 # results = soup.select("!!CSS 선택자로 지정해주세요 ex) td.subject > a") 선택자를 지정해야합니다.
# 필수 # articleUrl을 지정해야합니다.
# 선택 # file_name = nowDate+' filename.xlsx' 원하는 파일명의 형태로 지정해야합니다.

import urllib
import urllib.parse
import urllib.request
from bs4 import BeautifulSoup
import time
import requests
from openpyxl import Workbook
import datetime

url = "!!대상 URL입력해주세요"

response = urllib.request.urlopen(url)
soup = BeautifulSoup(response, 'html.parser')
results = soup.select("!!CSS 선택자로 지정해주세요 ex) td.subject > a")

fileTransDataTitle = list()
fileTransDataDes = list()

for result in results:
    print(str(len(results) - results.index(result)) +"sec...") 
    
    # path : result.attrs["href"]
    # result.attrs["href"]가 URL이 아닌 Path로 지정되어있을 때
    articleUrl = "!!대상 URL을 입력해주세요(url 변수와 다를경우)" + result.attrs["href"]
    articleUrlEncode = "http://" + urllib.parse.quote(articleUrl) # 경로(path)가 한글일때, 변환하여 적용하였습니다.
    
    fileTransDataTitle.append(result.text)
    fileTransDataDes.append(articleUrlEncode)

    time.sleep(1)

wb = Workbook()
sheet1 = wb.active

now = datetime.datetime.now()
nowDate = now.strftime('%Y-%m-%d')

file_name = nowDate+' filename.xlsx' # 년도-월-일자 filename.xlsx 으로 저장하였습니다.
sheet1.title = 'filename'

for i in range(1,11):
    sheet1.cell(row=i, column=1).value = i
    sheet1.cell(row=i, column=2).value = fileTransDataTitle[i]
    sheet1.cell(row=i, column=3).value = fileTransDataDes[i]

wb.save(filename=file_name)

print("success")
a = input("press Enter... exit") # 종료를 위해 적용하였습니다.

#pyinstaller을 사용하여 exe로 변환 후 사용 가능합니다.